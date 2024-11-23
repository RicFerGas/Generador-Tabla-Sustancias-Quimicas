# process_hds.py
import os
import sys
import json
import openai
import pandas as pd
from dotenv import load_dotenv
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,range_boundaries
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import pytesseract
import PyPDF2
import spacy
import fitz  # PyMuPDF
from rapidfuzz import fuzz, process
from schemas import HDSData


# Determine base path
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    base_path = sys._MEIPASS
else:
    base_path = os.path.dirname(os.path.abspath(__file__))
model_relative_path = os.path.join('models', 'xx_ent_wiki_sm', 'xx_ent_wiki_sm', 'xx_ent_wiki_sm-3.8.0')
model_path = os.path.join(base_path, model_relative_path)

# Load the model
nlp = spacy.load(model_path)

def is_valid_pdf(pdf_path: str) -> bool:
    """
    Validates if the PDF is structurally correct and can be opened.
    """
    if not os.path.isfile(pdf_path):
        print(f"File does not exist: {pdf_path}")
        return False

    try:
        # Check if it starts with %PDF-
        with open(pdf_path, 'rb') as file:
            header = file.read(4)
            if header != b'%PDF':
                print(f"File is not a valid PDF: {pdf_path}")
                return False

            # Attempt to open and read the PDF
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)

            # Check if there are any pages
            if num_pages == 0:
                print(f"PDF has no pages: {pdf_path}")
                return False
            
            # Check if any page has extractable text
            text_found = False
            for page in pdf_reader.pages:
                if page.extract_text():
                    text_found = True
                    break

            if not text_found:
                print(f"No extractable text in PDF: {pdf_path}")
                return False
            
            print(f"PDF is valid: {pdf_path} (pages: {num_pages})")
            return True
            
    except Exception as e:
        print(f"Invalid PDF {pdf_path}: {e}")
        return False
def validate_extracted_text(text: str, threshold: float = 0.6) -> bool:
    """
    Validates the extracted text from a PDF based on the word count threshold.
    """
    try:
        doc = nlp(text)
        word_count = sum(1 for token in doc if token.is_alpha)
        total_count = len(doc)
        print(f"word count: {word_count}, total count: {total_count}, rate: {word_count / total_count}")
        if total_count == 0:
            return False
        return (word_count / total_count) >= threshold
    except Exception as e:
        print(f"Error in text validation: {e}")
        return False

def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Extracts text from a PDF file, first using PyMuPDF and then OCR as a fallback.

    Args:
        pdf_path (str): The path to the PDF file.

    Returns:
        str: The extracted text from the PDF file.
    """
    text = ""

    # Intentar extraer texto usando PyMuPDF
    try:
        with fitz.open(pdf_path) as pdf:
            for page_num in range(len(pdf)):
                page = pdf.load_page(page_num)
                text += page.get_text()
    except Exception as e:
        print(f"Error extracting text from {pdf_path} using PyMuPDF: {e}")
        text = ""

    # Si se extrajo texto y pasa la validación, devolverlo
    if text and validate_extracted_text(text):
        return text
    else:
        print("No valid text found or extraction failed, proceeding to OCR.")

    # Intentar extraer texto usando OCR
    try:
        images = convert_from_path(pdf_path)
        extracted_text = ""
        for image in images:
            extracted_text += pytesseract.image_to_string(image, lang='spa')
        return extracted_text
    except Exception as e:
        print(f"Error during OCR process: {e}")
        return "OCR extraction failed."




def prompt_generation(hds: str) -> tuple:
    """
    Generates the system and user prompts for the OpenAI API based on the HDS text."""
    system_prompt = """
    Eres un asistente que extrae información de Hojas de Datos de Seguridad (HDS) para estudios de riesgo por sustancias químicas. 
    La información extraida debe estar siempre ** EN ESPAÑOL ** y seguir el formato especificado.
    Extrae la información relevante de la HDS proporcionada y devuélvela siguiendo el formato especificado, para el caso de las propiedades numericas,
    no pongas 0 si no ecnuentras la propiedad, el valor default es none."""
    user_prompt = f"La HDS es la siguiente:\n\n{hds}"
    return system_prompt, user_prompt

def get_completion_openai(system_prompt: str, user_prompt: str,
                          client ,model:str ="gpt-4o-2024-08-06",
                          max_tokens:int=4000):
    """
    Gets the completion from the OpenAI API using the system and user prompts.
    
    Args:
        system_prompt (str): The system prompt to provide context.
        user_prompt (str): The user prompt to generate the completion.
        client (openai.OpenAI): The OpenAI API client.
        model (str): The model to use for the completion.
        max_tokens (int): The maximum number of tokens to generate.

    Returns:
        HDSData: The parsed HDS data object from the completion.

    """
    
    completion = client.beta.chat.completions.parse(
    model=model,
    messages=[
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ],
    # Pasamos el modelo Pydantic directamente
    response_format=HDSData,
    max_tokens=max_tokens,
    temperature=0.0,
    top_p=1,
)

# Obtenemos el objeto parseado directamente
    hds_data = completion.choices[0].message.parsed
    # print(hds_data.model_dump())
    return hds_data


def flatten_hds_data(hds_dict: dict, retc_data: dict, gei_data: dict) -> list:
    """
    Flattens the HDS data into a list of dictionaries for Excel export.

    Args:
        hds_dict (dict): The HDS data dictionary.
        retc_data (dict): The RETC data dictionary.
        gei_data (dict): The GEI data dictionary.

    Returns:
        list: A list of dictionaries with flattened data for each substance.
    """
    sustancias_flat = []

    # Información común para todas las filas de la sustancia
    pictogram_columns = {
    'bomba_explotando': 'Explosivos',
    'llama': 'Inflamables',
    'llama_sobre_circulo': 'Comburentes',
    'cilindro_de_gas': 'Gases Comprimidos',
    'corrosion': 'Corrosivos',
    'calavera_tibias_cruzadas': 'Toxicidad Aguda',
    'signo_de_exclamacion': 'Irritantes',
    'peligro_para_la_salud': 'Peligro Crónico para la Salud',
    'medio_ambiente': 'Peligro Ambiental'}
    
    

    common_info = {
        'Archivo': hds_dict.get('Archivo'),
        'Nombre de la Sustancia Química': hds_dict.get('nombre_sustancia_quimica'),
        'Idioma de la HDS': hds_dict.get('idioma'),
        'Estado Físico': hds_dict.get('estado_fisico'),
        'Sujeta a RETC': hds_dict.get('sujeta_retc'),
        'Sujeta a GEI': hds_dict.get('sujeta_gei'),
        'pH de la sustancia': hds_dict.get('ph'),
        'Palabra de Advertencia': hds_dict.get('palabra_advertencia'),
        'Indicaciones de toxicología': hds_dict.get('indicaciones_toxicologia'),
        'Pictogramas': ', '.join(hds_dict.get('pictogramas', [])),
        'Olor': hds_dict.get('olor'),
        'Color': hds_dict.get('color'),
        'Propiedades Explosivas': hds_dict.get('propiedades_explosivas'),
        'Propiedades Comburentes': hds_dict.get('propiedades_comburentes'),
        'Tamaño de Partícula': hds_dict.get('tamano_particula'),
    }
    # Extract pictogram booleans and map to hazards
    pictogramas = hds_dict.get('pictogramas', {})
    for field_name, hazard_name in pictogram_columns.items():
        is_present = pictogramas.get(field_name, False)
        common_info[hazard_name] = is_present

    # Desanidar propiedades numéricas
    propiedades = [
        'temperatura_ebullicion',
        'punto_congelacion',
        'densidad',
        'punto_inflamacion',
        'solubilidad_agua',
        'velocidad_evaporacion',
        'presion_vapor'
    ]
    for prop in propiedades:
        propiedad = hds_dict.get(prop)
        if propiedad:
            common_info[f'{prop} (Valor)'] = propiedad.get('valor')
            common_info[f'{prop} (Unidades)'] = propiedad.get('unidades')
        else:
            common_info[f'{prop} (Valor)'] = None
            common_info[f'{prop} (Unidades)'] = None

    # Límites de inflamabilidad
    common_info['Límite inferior de inflamabilidad'] = hds_dict.get('limite_inf_inflamabilidad')
    common_info['Límite superior de inflamabilidad'] = hds_dict.get('limite_sup_inflamabilidad')

    # Manejar componentes
    componentes = hds_dict.get('componentes', [])
    if componentes:
        for componente in componentes:
            row = common_info.copy()
            row['Nombre del Componente'] = componente.get('nombre')
            row['Número CAS del Componente'] = componente.get('numero_cas')
            row['Porcentaje del Componente'] = componente.get('porcentaje')

            # Verificar si el número CAS está en la tabla RETC
            cas_number = componente.get('numero_cas')
            nombre_componente = componente.get('nombre')

            # Buscar en RETC por número CAS
            retc_entry = retc_data.get(cas_number)

            # Si no se encuentra por CAS, intentar buscar por nombre usando coincidencia difusa
            if not retc_entry and nombre_componente:
                # Crear una lista de nombres comunes de RETC
                retc_nombres = [entry['componente_retc'] for entry in retc_data.values()]
                # Realizar coincidencia difusa
                match = process.extractOne(nombre_componente, 
                                           retc_nombres, 
                                           scorer=fuzz.token_sort_ratio)
                if match and match[1] >= 80:  # Umbral de similitud del 80%
                    # Obtener la entrada correspondiente
                    for entry in retc_data.values():
                        if entry['componente_retc'] == match[0]:
                            retc_entry = entry
                            break

            # Asignar datos de RETC si se encontró una coincidencia
            if retc_entry:
                row['Componente RETC'] = retc_entry.get('componente_retc')
                row['MPU'] = retc_entry.get('mpu')
                row['Emision Transferencia'] = retc_entry.get('emision_transferencia')
            else:
                row['Componente RETC'] = None
                row['MPU'] = None
                row['Emision Transferencia'] = None

            # Hacer lo mismo para GEI
            # Buscar en GEI por número CAS
            gei_entry = gei_data.get(cas_number)

            # Si no se encuentra por CAS, intentar buscar por nombre usando coincidencia difusa
            if not gei_entry and nombre_componente:
                gei_nombres = [entry['nombre_comun'] for entry in gei_data.values()]
                match = process.extractOne(nombre_componente,
                                            gei_nombres,
                                            scorer=fuzz.token_sort_ratio)
                if match and match[1] >= 90:
                    print(f"GEI match between {nombre_componente} and {match[0]}: {match[1]}")
                    for entry in gei_data.values():
                        if entry['nombre_comun'] == match[0]:
                            gei_entry = entry
                            break

            if gei_entry:
                row['Componente GEI'] = gei_entry.get('nombre_comun')
                row['Potencial de Calentamiento Global'] = gei_entry.get('PCG')
            else:
                row['Componente GEI'] = None
                row['Potencial de Calentamiento Global'] = None

            sustancias_flat.append(row)
    else:
        # Si no hay componentes, agregar la sustancia sin detalles de componentes
        row = common_info.copy()
        # Agregar campos vacíos
        row['Nombre del Componente'] = None
        row['Número CAS del Componente'] = None
        row['Porcentaje del Componente'] = None
        row['Componente RETC'] = None
        row['MPU'] = None
        row['Emision Transferencia'] = None
        row['Componente GEI'] = None
        row['Potencial de Calentamiento Global'] = None
        sustancias_flat.append(row)

    # Manejar Indicaciones de peligro H y Consejos de Prudencia P
    indicaciones_h = hds_dict.get('identificaciones_peligro_h', [])
    consejos_p = hds_dict.get('consejos_prudencia_p', [])

    # Concatenar códigos y descripciones
    common_info['Identificaciones de peligro H'] = '; '.join(
        [f"{i['codigo']}: {i['descripcion']}" for i in indicaciones_h])
    common_info['Consejos de Prudencia P'] = '; '.join(
        [f"{i['codigo']}: {i['descripcion']}" for i in consejos_p])

    # Actualizar las filas existentes con las indicaciones
    for row in sustancias_flat:
        row.update({
            'Identificaciones de peligro H': common_info['Identificaciones de peligro H'],
            'Consejos de Prudencia P': common_info['Consejos de Prudencia P']
        })

    return sustancias_flat


def combine_headers(ws, df):
    """
    Combines cells in the header row (row 1) to organize related properties.
    """
    # Define the column groups and their respective combined headers
    header_groups = {
        'Pictogramas': ['Explosivos', 'Inflamables', 'Comburentes', 'Gases Comprimidos', 
                        'Corrosivos', 'Toxicidad Aguda', 'Irritantes', 
                        'Peligro Crónico para la Salud', 'Peligro Ambiental'],
        'Propiedades Físicas': ['temperatura_ebullicion (Valor)', 'temperatura_ebullicion (Unidades)', 
                                'punto_congelacion (Valor)', 'punto_congelacion (Unidades)', 
                                'densidad (Valor)', 'densidad (Unidades)', 
                                'punto_inflamacion (Valor)', 'punto_inflamacion (Unidades)', 
                                'presion_vapor (Valor)', 'presion_vapor (Unidades)', 
                                'solubilidad_agua (Valor)', 'solubilidad_agua (Unidades)']
    }

    for group_name, columns in header_groups.items():
        # Find the columns by name in the DataFrame
        column_indices = [df.columns.get_loc(col) + 1 for col in columns if col in df.columns]

        if column_indices:
            # Combine the cells in the first row
            first_col = min(column_indices)
            last_col = max(column_indices)
            ws.merge_cells(start_row=1, start_column=first_col, end_row=1, end_column=last_col)
            header_cell = ws.cell(row=1, column=first_col)
            header_cell.value = group_name
            header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust the individual headers in row 2
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num)
        # Skip merged cells and only set values for the top-left cell of a merged range
        if not isinstance(cell, MergedCell):
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center')

def export_to_excel(data:list, peligro_a_pictograma:dict, output_file:str ="output.xlsx")->None:
    """
    Exporta los datos de las sustancias a un archivo Excel con celdas combinadas y alineadas.

    Args:
        data (list): La lista de datos de las sustancias.
        output_file (str): La ruta del archivo de salida Excel.


    """

    # Crear DataFrame de pandas
    df = pd.DataFrame(data)


    # Define the desired order of columns
    column_order = [
        'Archivo',                         # 1. File name
        'Nombre de la Sustancia Química',   # 2. Substance name
        'Idioma de la HDS',                 # 3. Language
        'Sujeta a GEI',                     # 5. GEI component
        'Sujeta a RETC',
        'Nombre del Componente',            # 4. Component name
        'Número CAS del Componente',        # 7. Component CAS number
        'Porcentaje del Componente',        # 8. Component percentage
        'Componente GEI',                   # 9. GEI component
        'Potencial de Calentamiento Global',# 10. Global Warming Potential
        'Componente RETC',                  # 11. RETC component
        'MPU',                              # 12. MPU
        'Emision Transferencia',            # 13. Emission transfer
        'Palabra de Advertencia',           # 14. Warning word
        'Indicaciones de toxicología',      # 15. Toxicology indications
        'Explosivos',                       # 16. Explosives
        'Inflamables',                      # 17. Flammable
        'Comburentes',                      # 18. Oxidizing
        'Gases Comprimidos',                # 19. Compressed gases
        'Corrosivos',                       # 20. Corrosive
        'Toxicidad Aguda',                  # 21. Acute toxicity
        'Irritantes',                       # 22. Irritants
        'Peligro Crónico para la Salud',    # 23. Chronic health hazard
        'Peligro Ambiental',                # 24. Environmental hazard
        'Estado Físico',                    # 25. Physical state
        'Identificaciones de peligro H',    # 26. Hazard identifications H
        'Consejos de Prudencia P',          # 27. Precautionary statements P
        'Olor',                             # 28. Odor
        'Color',                            # 29. Color
        'pH de la sustancia',               # 30. pH
        'temperatura_ebullicion (Valor)',   # 31. Boiling temperature (Value)
        'temperatura_ebullicion (Unidades)',# 32. Boiling temperature (Units)
        'punto_congelacion (Valor)',        # 33. Freezing point (Value)
        'punto_congelacion (Unidades)',     # 34. Freezing point (Units)
        'densidad (Valor)',                 # 35. Density (Value)
        'densidad (Unidades)',              # 36. Density (Units)
        'punto_inflamacion (Valor)',        # 37. Flash point (Value)
        'punto_inflamacion (Unidades)',     # 38. Flash point (Units)
        'velocidad_evaporacion (Valor)',    # 39. Evaporation rate (Value)
        'velocidad_evaporacion (Unidades)', # 40. Evaporation rate (Units)
        'presion_vapor (Valor)',            # 41. Vapor pressure (Value)
        'presion_vapor (Unidades)',         # 42. Vapor pressure (Units)
        'solubilidad_agua (Valor)',         # 43. Water solubility (Value)
        'solubilidad_agua (Unidades)',      # 44. Water solubility (Units)
        'Propiedades Explosivas',           # 45. Explosive properties
        'Propiedades Comburentes',          # 46. Oxidizing properties
        'Tamaño de Partícula',              # 47. Particle size
        'Límite inferior de inflamabilidad',# 48. Lower flammability limit
        'Límite superior de inflamabilidad' # 49. Upper flammability limit
    ]

    # Reorder the DataFrame columns
    df = df[column_order]
    

    # Save the DataFrame to a temporary Excel file
    temp_file = "temp_output.xlsx"
    df.to_excel(temp_file, index=False)

    # Load the Excel file with openpyxl for customization
    wb = load_workbook(temp_file)
    ws = wb.active
    combine_headers(ws, df)
    # Insert a new row at the top for the merged cell "Pictogramas"
    ws.insert_rows(1)
    # Merge cells for property columns
    propiedades = [
        ('temperatura_ebullicion (Valor)', 'temperatura_ebullicion (Unidades)', 'Temperatura de Ebullición'),
        ('punto_congelacion (Valor)', 'punto_congelacion (Unidades)', 'Punto de Congelación'),
        ('densidad (Valor)', 'densidad (Unidades)', 'Densidad'),
        ('punto_inflamacion (Valor)', 'punto_inflamacion (Unidades)', 'Punto de Inflamación'),
        ('velocidad_evaporacion (Valor)', 'velocidad_evaporacion (Unidades)', 'Velocidad de Evaporación'),
        ('presion_vapor (Valor)', 'presion_vapor (Unidades)', 'Presión de Vapor'),
        ('solubilidad_agua (Valor)', 'solubilidad_agua (Unidades)', 'Solubilidad en Agua')
    ]

    # First, create a set of all property columns for easy checking
    propiedad_columns = set()
    for valor_col, unidades_col, propiedad_name in propiedades:
        propiedad_columns.update([valor_col, unidades_col])

   # Crear un conjunto de celdas combinadas en la hoja
    # Modificación para manejar celdas combinadas
    merged_ranges = ws.merged_cells.ranges
    merged_cells = {
        (row, col)
        for merged_range in merged_ranges
        for row in range(merged_range.min_row, merged_range.max_row + 1)
        for col in range(merged_range.min_col, merged_range.max_col + 1)
    }

    for col_num, column_title in enumerate(df.columns, 1):
        if (2, col_num) not in merged_cells:  # Si la celda no está combinada
            ws.cell(row=2, column=col_num).value = column_title
        elif any(
            merged_range.min_row == 2 and col_num == merged_range.min_col
            for merged_range in merged_ranges
        ):  # Si es la celda principal de un rango combinado
            ws.cell(row=2, column=col_num).value = column_title
    # Now handle the propiedades columns
    for valor_col, unidades_col, propiedad_name in propiedades:
        # Get column indices
        valor_idx = df.columns.get_loc(valor_col) + 1
        unidades_idx = df.columns.get_loc(unidades_col) + 1

        # Merge the property name across the two columns
        ws.merge_cells(start_row=1, start_column=valor_idx, end_row=1, end_column=unidades_idx)
        propiedad_cell = ws.cell(row=1, column=valor_idx)
        propiedad_cell.value = propiedad_name
        propiedad_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set 'Valor' and 'Unidades' in the second row
        ws.cell(row=2, column=valor_idx).value = 'Valor'
        ws.cell(row=2, column=valor_idx).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=unidades_idx).value = 'Unidades'
        ws.cell(row=2, column=unidades_idx).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust row heights for the header rows
    ws.row_dimensions[1].height = 30  # Adjust as needed
    ws.row_dimensions[2].height = 20  # Adjust as needed
    # Define the columns that correspond to the pictograms
    pictogram_columns = [
        'Explosivos',
        'Inflamables',
        'Comburentes',
        'Gases Comprimidos',
        'Corrosivos',
        'Toxicidad Aguda',
        'Irritantes',
        'Peligro Crónico para la Salud',
        'Peligro Ambiental',
    ]
    # Reemplazar True por 'X' y False por '' en las columnas de pictogramas
    df[pictogram_columns] = df[pictogram_columns].replace({True: 'X', False: ''})
    # Find the column indices for the pictogram columns
    pictogram_col_indices = [df.columns.get_loc(col) + 1 for col in pictogram_columns]

    # Merge cells above the pictogram columns
    first_pictogram_col = min(pictogram_col_indices)
    last_pictogram_col = max(pictogram_col_indices)
    ws.merge_cells(
        start_row=1,
        start_column=first_pictogram_col,
        end_row=1,
        end_column=last_pictogram_col
    )

    # Write "Pictogramas" in the merged cell
    pictogram_header_cell = ws.cell(row=1, column=first_pictogram_col)
    pictogram_header_cell.value = "Pictogramas"
    pictogram_header_cell.alignment = Alignment(horizontal='center', vertical='center')

    
    # Obtener las columnas que ya tienen 'Valor' y 'Unidades' asignados
    valor_unidades_indices = []
    for valor_col, unidades_col, _ in propiedades:
        valor_idx = df.columns.get_loc(valor_col) + 1
        unidades_idx = df.columns.get_loc(unidades_col) + 1
        valor_unidades_indices.extend([valor_idx, unidades_idx])

    # Ajustar los encabezados hacia abajo por la fila insertada, sin sobrescribir 'Valor' y 'Unidades'
    for col_num, column_title in enumerate(df.columns, 1):
        if col_num not in valor_unidades_indices:
            ws.cell(row=2, column=col_num).value = column_title
            ws.cell(row=2, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

    # Insertar imágenes en los encabezados de las columnas de pictogramas
    for col_index in pictogram_col_indices:
        col_letter = get_column_letter(col_index)
        hazard_name = ws.cell(row=2, column=col_index).value

        # Obtener la información del pictograma desde el JSON
        pictograma_info = peligro_a_pictograma.get(hazard_name)
        if pictograma_info:
            image_path = pictograma_info['ruta_imagen']
            # Asegurarse de que la ruta sea absoluta
            image_path = os.path.abspath(image_path)
            if os.path.exists(image_path):
                
                img = Image(image_path)
                img.height = 40  # Ajusta el tamaño según sea necesario
                img.width = 40
                img.anchor = f"{col_letter}2"
                ws.add_image(img)
                # Limpiar el texto del encabezado
                ws.cell(row=2, column=col_index).value = None
            else:
                print(f"No se encontró la imagen para {hazard_name} en {image_path}")
        else:
            print(f"No hay información de pictograma para el peligro: {hazard_name}")

        # Adjust row heights for the header rows
    ws.row_dimensions[1].height = 30  # Height for "Pictogramas"
    ws.row_dimensions[2].height = 50  # Height for images

    # Agregar las columnas de 'Valor' y 'Unidades' a propiedad_columns
    for valor_col, unidades_col, _ in propiedades:
        propiedad_columns.update([valor_col, unidades_col])

    # Definir las columnas que NO son propiedades ni pictogramas
    non_property_pictogram_columns = [
        col for col in df.columns if col not in propiedad_columns and col not in pictogram_columns
    ]

    # Inicializar diccionarios para el seguimiento de valores actuales y filas de inicio
    current_values = {col: None for col in non_property_pictogram_columns}
    start_rows = {col: 3 for col in non_property_pictogram_columns}  # Comienza en la fila 3
    current_sustancia = None

    for row in range(3, ws.max_row + 1):
        sustancia_value = ws[f'B{row}'].value  # Columna 'Nombre de la Sustancia Química'
        if sustancia_value != current_sustancia:
            # Si cambia la sustancia, combina celdas anteriores
            for col_name in non_property_pictogram_columns:
                if current_values[col_name] is not None:
                    col_idx = df.columns.get_loc(col_name) + 1
                    ws.merge_cells(start_row=start_rows[col_name], start_column=col_idx,
                                   end_row=row - 1, end_column=col_idx)
                current_values[col_name] = ws.cell(row=row, column=df.columns.get_loc(col_name) + 1).value
                start_rows[col_name] = row
            current_sustancia = sustancia_value
        else:
            # Si la sustancia es la misma, verifica para combinar celdas
            for col_name in non_property_pictogram_columns:
                col_idx = df.columns.get_loc(col_name) + 1
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value == current_values[col_name]:
                    ws.cell(row=row, column=col_idx).value = None  # Limpiar celdas duplicadas
                else:
                    # Combina celdas anteriores
                    if current_values[col_name] is not None:
                        ws.merge_cells(start_row=start_rows[col_name], start_column=col_idx,
                                       end_row=row - 1, end_column=col_idx)
                    current_values[col_name] = cell_value
                    start_rows[col_name] = row

    # Combinar celdas para el último grupo
    for col_name in non_property_pictogram_columns:
        if current_values[col_name] is not None:
            col_idx = df.columns.get_loc(col_name) + 1
            ws.merge_cells(start_row=start_rows[col_name], start_column=col_idx,
                           end_row=ws.max_row, end_column=col_idx)

    # Crear un mapeo de nombres de columna a letras de Excel
    column_letters = {}
    for idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        column_letters[col_name] = col_letter

    # Establecer los anchos de columna
    fixed_column_widths = {
        'Archivo': 10,
        'Nombre de la Sustancia Química': 15,
        'Idioma de la HDS': 10,
        'Sujeta a GEI': 8,
        'Sujeta a RETC': 8,
        'Nombre del Componente': 15,
        'Número CAS del Componente': 12,
        'Porcentaje del Componente': 12,
        'Componente GEI': 15,
        'Potencial de Calentamiento Global': 8,
        'Componente RETC': 15,
        'MPU': 8,
        'Emision Transferencia': 8,
        'Palabra de Advertencia': 10,
        'Indicaciones de toxicología': 20,
        'Estado Físico': 8,
        'Identificaciones de peligro H': 20,
        'Consejos de Prudencia P': 20,
        'Olor': 10,
        'Color': 10,
        'pH de la sustancia': 8,
        'Propiedades Explosivas': 20,
        'Propiedades Comburentes': 20,
        'Tamaño de Partícula': 8,
        'Límite inferior de inflamabilidad': 8,
        'Límite superior de inflamabilidad': 8,
        'Explosivos': 5,
        'Inflamables': 5,
        'Comburentes': 5,
        'Gases Comprimidos': 5,
        'Corrosivos': 5,
        'Toxicidad Aguda': 5,
        'Irritantes': 5,
        'Peligro Crónico para la Salud': 5,
        'Peligro Ambiental': 5,
        'Valor': 10,  # Añade los anchos para las columnas de propiedades y pictogramas si es necesario
        'Unidades': 8,
        # Añade los anchos para las columnas de propiedades y pictogramas si es necesario
    }

    for col_name, width in fixed_column_widths.items():
        if col_name in column_letters:
            col_letter = column_letters[col_name]
            ws.column_dimensions[col_letter].width = width

    # Ajustar la altura de las filas para que se vea todo el texto
    for row in ws.iter_rows(min_row=3, max_col=ws.max_column, max_row=ws.max_row):
        max_height = 1
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')  # Ajuste automático del texto y centrado
            if cell.value and isinstance(cell.value, str):
                max_height = max(max_height, len(cell.value.split('\n')))
        ws.row_dimensions[cell.row].height = max_height * 15  # Ajuste según la cantidad de líneas

    # Guardar el archivo final
    wb.save(output_file)
    # Eliminar el archivo temporal
    os.remove(temp_file)


def process_hds_folder(folder_path:str, project_name:str, client, excel_output:str, json_output:str, progress_callback=None):
    """
    Procesa una carpeta de archivos PDF de Hojas de Datos de Seguridad (HDS) y genera un archivo Excel con los datos extraídos.

    Args:
        folder_path (str): La ruta de la carpeta que contiene los archivos PDF.
        project_name (str): El nombre del proyecto o carpeta.
        client (openai.OpenAI): El cliente de OpenAI API.
        excel_output (str): La ruta del archivo Excel de salida.
        json_output (str): La ruta del archivo JSON de salida.
        progress_callback (function): Una función de callback para informar del progreso.

    Returns:
        tuple: Una tupla con una lista de errores y una lista de todas las sustancias procesadas.
    """
    # Cargar la tabla RETC
    # Determine base path
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    

    retc_file_path = os.path.join(base_path,'data_sets', 'retc_table.json')
    print("RETC FILE PATH: ",retc_file_path)

    gei_file_path = os.path.join(base_path,'data_sets', 'gei_table.json')
    print("GEI FILE PATH: ",gei_file_path)

    with open(retc_file_path, 'r', encoding='utf-8') as f:
        retc_list = json.load(f)
        # Convertir la lista de dicts a un dict con el número CAS como clave para acceso rápido
        retc_data = {entry['numero_cas']: entry for entry in retc_list}
    # Cargar el mapeo de pictogramas desde el archivo JSON
    # Ruta al archivo JSON de pictogramas
    pictogram_json_path = os.path.join(base_path, 'data_sets', 'pictograms_mapping.json')

    # Cargar el JSON de pictogramas
    with open(pictogram_json_path, 'r', encoding='utf-8') as f:
        pictogram_data = json.load(f)
    # Crear un diccionario que mapee el peligro al pictograma completo
    peligro_a_pictograma = {item['peligro']: item for item in pictogram_data}

    with open(gei_file_path, 'r', encoding='utf-8') as f:
        gei_list = json.load(f)
        # Crear un diccionario con número CAS como clave
        gei_data_cas = {entry['CAS']: entry for entry in gei_list}
        # Crear un diccionario con nombre común como clave
        gei_data_nombre = {entry['nombre_comun']: entry for entry in gei_list}
        # Combinar ambos
        gei_data = {**gei_data_cas, **gei_data_nombre}
    all_sustancias = []
    errores = []
    sustancias_data = []  # Lista para almacenar los datos procesados para Excel
    progress_percent = 0

    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
    total_files = len(pdf_files)
    # progress_callback(progress_percent, pdf_files[0] if pdf_files else "")

    for index, filename in enumerate(pdf_files):
        
        
        file_path = os.path.join(folder_path, filename)
         # Calcular el progreso y llamar al callback
        if progress_callback:
            progress_callback(progress_percent, filename)

        # Leer el contenido de la HDS
        hds_text = extract_text_from_pdf(file_path)
        # Generar el prompt
        system_prompt, user_prompt = prompt_generation(hds_text)

        try:
            # Obtener el objeto HDSData directamente del LLM
            hds_data = get_completion_openai(system_prompt, user_prompt, client)

            # Convertir el objeto HDSData a dict
            hds_dict = hds_data.dict()
            hds_dict["Archivo"] = filename

            # Postprocesar para Excel
            sustancias_data.extend(flatten_hds_data(hds_dict, retc_data,gei_data))

            print(f"Procesado correctamente: {filename}")
            progress_percent = (index + 1) * 100 // total_files
            all_sustancias.append(hds_dict)
        except Exception as e:
            print(f"Error al procesar {filename}: {e}")
            errores.append(filename)


    # Guardar el JSON final con todas las sustancias
    with open(json_output, 'w', encoding='utf-8') as output_file:
        json.dump(all_sustancias, output_file, indent=2, ensure_ascii=False)

    # Generar el Excel formateado con celdas combinadas y alineadas
    export_to_excel(sustancias_data,peligro_a_pictograma,excel_output)

    print(f"Proceso completado. Los archivos {json_output} y {excel_output} han sido generados.")
    return errores, all_sustancias

if __name__ == "__main__":
    load_dotenv()
    openai.api_key = os.getenv("OPENAI_API_KEY")
    example_path="ejemplo/mini_test"
    client = openai.OpenAI()
    process_hds_folder(example_path, "Prueba GEI y RETC",client,'ejemplo/outputs/prueba_pictogramas1.xlsx','ejemplo/outputs/prueba_pictogramas1')
