# excel_postprocess.py
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from rapidfuzz import fuzz, process
import pandas as pd
# generador_tabla_sustq.py
import pandas as pd
import json
from datetime import datetime
from pathlib import Path
from rapidfuzz import fuzz, process
import unicodedata

class GeneradorTablaSustQ:
    """
    Clase para generar una tabla de sustancias químicas a partir de un diccionario de datos de HDS
    """
    def __init__(self, config_path: str = None):
        """Initialize with paths to reference data files"""

        if config_path is None:
            config_path = 'data_sets'
            
        # Convert to Path object and resolve data_sets subdirectory
        self.config_path = Path(config_path)
        if self.config_path.name != 'data_sets':
            self.config_path = self.config_path / 'data_sets'
            
        print(f"Config path initialized as: {self.config_path}")
        self._load_reference_data()
    
    def _load_reference_data(self):
        """Load RETC, GEI and pictogram reference data"""
        # Load RETC data
        retc_list = self._load_json_file("retc_table.json")
        self.retc_data = {entry['numero_cas']: entry for entry in retc_list}
        
        # Load pictogram data
        pictogram_data = self._load_json_file("pictograms_mapping.json")
        self.peligro_a_pictograma = {entry['peligro']: entry for entry in pictogram_data}
        
        # Load GEI data
        gei_list = self._load_json_file("gei_table.json")
        self.gei_data = {entry['CAS']: entry for entry in gei_list}
    
    def _load_json_file(self, filename: str) -> dict:
        """Helper to load JSON files"""
        file_path = self.config_path / filename
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    

    def flatten_hds_data(self, hds_data: list) -> list:
        """
        Flattens the HDS data into a list of dictionaries for Excel export.

        Args:
            hds_dict (dict): The HDS data dictionary.

        Returns:
            list: A list of dictionaries with flattened data for each substance.
        """
        sustancias_flat = []

        for hds_dict in hds_data:
            common_info = self._extract_common_info(hds_dict)
            self._extract_pictogram_info(hds_dict, common_info)
            self._extract_propiedades_info(hds_dict, common_info)
            self._extract_indicaciones_consejos_info(hds_dict, common_info)
            self._extract_valores_limite(hds_dict, common_info)
            self._calculate_riesgo_tabla_5(common_info)
            self._calculate_riesgo_tabla_6(hds_dict, common_info)
            self._determine_volatilidad(hds_dict, common_info)
            self._extract_componentes_info(hds_dict, common_info, sustancias_flat)
            
           

        return sustancias_flat

    def _extract_common_info(self, hds_dict: dict) -> dict:
        """Extract common information for all rows of the substance."""
        estado_fisico_data = hds_dict.get('estado_fisico', {})
        if estado_fisico_data.get('solido_baja') or estado_fisico_data.get('solido_media') or estado_fisico_data.get('solido_alta'):
            estado_fisico = 'Sólido'
        elif estado_fisico_data.get('liquido'):
            estado_fisico = 'Líquido'
        elif estado_fisico_data.get('gaseoso'):
            estado_fisico = 'Gaseoso'
        else:
            estado_fisico = 'No determinado'
        return {
            'Archivo': hds_dict.get('Archivo'),
            'Nombre de la Sustancia Química': hds_dict.get('nombre_sustancia_quimica'),
            'Idioma de la HDS': hds_dict.get('idioma'),
            'Estado Físico': estado_fisico,
            'Sujeta a RETC': hds_dict.get('sujeta_retc'),
            'Sujeta a GEI': hds_dict.get('sujeta_gei'),
            'pH de la sustancia': hds_dict.get('ph'),
            'Palabra de Advertencia': hds_dict.get('palabra_advertencia'),
            'Indicaciones de toxicología': hds_dict.get('indicaciones_toxicologia'),
            'Pictogramas': ', '.join(hds_dict.get('pictogramas', [])),
            'Olor': hds_dict.get('olor'),
            'Color': hds_dict.get('color'),
            'Propiedades Explosivas': hds_dict.get('propiedades_explosivas'),
            'Propiedades Comburentes': hds_dict.get('propiedades_comburentes'),
            'Tamaño de Partícula': hds_dict.get('tamano_particula'),
        }

    def _extract_pictogram_info(self, hds_dict: dict, common_info: dict) -> None:
        """Extract pictogram information and map to hazards."""
        pictogram_columns = {
            'bomba_explotando': 'Pictogramas Explosivos',
            'llama': 'Pictogramas Inflamables',
            'llama_sobre_circulo': 'Pictogramas Comburentes',
            'cilindro_de_gas': 'Pictogramas Gases Comprimidos',
            'corrosion': 'Pictogramas Corrosivos',
            'calavera_tibias_cruzadas': 'Pictogramas Toxicidad Aguda',
            'signo_de_exclamacion': 'Pictogramas Irritantes',
            'peligro_para_la_salud': 'Pictogramas Peligro Crónico para la Salud',
            'medio_ambiente': 'Pictogramas Peligro Ambiental'
        }

        pictogramas = hds_dict.get('pictogramas', {})
        for field_name, hazard_name in pictogram_columns.items():
            is_present = pictogramas.get(field_name, False)
            common_info[hazard_name] = 'X' if is_present else None 

    def _extract_propiedades_info(self, hds_dict: dict, common_info: dict) -> None:
        """Extract numerical properties information."""
        propiedades = [
            'temperatura_ebullicion',
            'punto_congelacion',
            'densidad',
            'punto_inflamacion',
            # 'solubilidad_agua',
            # 'velocidad_evaporacion',
            'presion_vapor'
        ]
        for prop in propiedades:
            propiedad = hds_dict.get(prop)
            if propiedad:
                common_info[f'{prop} (Valor)'] = propiedad.get('valor')
                common_info[f'{prop} (Unidades)'] = propiedad.get('unidades')
            else:
                common_info[f'{prop} (Valor)'] = None
                common_info[f'{prop} (Unidades)'] = None

        common_info['Límite inferior de inflamabilidad'] = hds_dict.get('limite_inf_inflamabilidad')
        common_info['Límite superior de inflamabilidad'] = hds_dict.get('limite_sup_inflamabilidad')
        velocidad = hds_dict.get('velocidad_evaporacion', {})
        solubilidad = hds_dict.get('solubilidad_agua', {})
        peso_molecular = hds_dict.get('peso_molecular', {})
        
        common_info["Velocidad de Evaporación Valor"] = velocidad.get('valor') if velocidad else None
        common_info["Velocidad de Evaporación Unidades"] = velocidad.get('unidades') if velocidad else None
        common_info["Solubilidad en Agua Valor"] = solubilidad.get('valor') if solubilidad else None
        common_info["Solubilidad en Agua Unidades"] = solubilidad.get('unidades') if solubilidad else None
        common_info["Peso Molecular Valor"] = peso_molecular.get('valor') if peso_molecular else None
        common_info["Peso Molecular Unidades"] = peso_molecular.get('unidades') if peso_molecular else None


    def _extract_componentes_info(self, hds_dict: dict, common_info: dict, sustancias_flat: list) -> None:
        """Extract components information."""
        componentes = hds_dict.get('componentes', [])
        if componentes:
            for componente in componentes:
                row = common_info.copy()
                row['Nombre del Componente'] = componente.get('nombre')
                row['Número CAS del Componente'] = componente.get('numero_cas')
                row['Porcentaje del Componente'] = componente.get('porcentaje')

                cas_number = componente.get('numero_cas')
                nombre_componente = componente.get('nombre')

                retc_entry = self._find_retc_entry(cas_number, nombre_componente)
                if retc_entry:
                    row['Componente RETC'] = retc_entry.get('componente_retc')
                    row['MPU'] = retc_entry.get('mpu')
                    row['Emision Transferencia'] = retc_entry.get('emision_transferencia')
                else:
                    row['Componente RETC'] = None
                    row['MPU'] = None
                    row['Emision Transferencia'] = None

                gei_entry = self._find_gei_entry(cas_number, nombre_componente)
                if gei_entry:
                    row['Componente GEI'] = gei_entry.get('nombre_comun')
                    row['Potencial de Calentamiento Global'] = gei_entry.get('PCG')
                else:
                    row['Componente GEI'] = None
                    row['Potencial de Calentamiento Global'] = None

                sustancias_flat.append(row)
        else:
            row = common_info.copy()
            row['Nombre del Componente'] = None
            row['Número CAS del Componente'] = None
            row['Porcentaje del Componente'] = None
            row['Componente RETC'] = None
            row['MPU'] = None
            row['Emision Transferencia'] = None
            row['Componente GEI'] = None
            row['Potencial de Calentamiento Global'] = None
            sustancias_flat.append(row)

    def _extract_indicaciones_consejos_info(self, hds_dict: dict, common_info: dict) -> None:
        """Extract hazard indications and precautionary statements."""
        indicaciones_h = hds_dict.get('identificaciones_peligro_h', [])
        consejos_p = hds_dict.get('consejos_prudencia_p', [])

        common_info['Identificaciones de peligro H'] = '; '.join(
            [f"{i['codigo']}: {i['descripcion']}" for i in indicaciones_h])
        common_info['Consejos de Prudencia P'] = '; '.join(
            [f"{i['codigo']}: {i['descripcion']}" for i in consejos_p])


    def _find_retc_entry(self, cas_number: str, nombre_componente: str) -> dict:
        """Find RETC entry by CAS number or component name."""
        retc_entry = self.retc_data.get(cas_number)
        if not retc_entry and nombre_componente:
            retc_nombres = [entry['componente_retc'] for entry in self.retc_data.values()]
            match = process.extractOne(nombre_componente, retc_nombres, scorer=fuzz.token_sort_ratio)
            if match and match[1] >= 95:
                for entry in self.retc_data.values():
                    if entry['componente_retc'] == match[0]:
                        retc_entry = entry
                        break
        return retc_entry

    def _find_gei_entry(self, cas_number: str, nombre_componente: str) -> dict:
        """Find GEI entry by CAS number or component name."""
        gei_entry = self.gei_data.get(cas_number)
        if not gei_entry and nombre_componente:
            gei_nombres = [entry['nombre_comun'] for entry in self.gei_data.values()]
            match = process.extractOne(nombre_componente, gei_nombres, scorer=fuzz.token_sort_ratio)
            if match and match[1] >= 95:
                for entry in self.gei_data.values():
                    if entry['nombre_comun'] == match[0]:
                        gei_entry = entry
                        break
        return gei_entry
    def _extract_valores_limite(self, hds_dict: dict, common_info: dict) -> None:
        """Extract exposure limits into separate columns with numeric values and units."""
        routes = ['oral', 'inhalacion', 'cutanea']
        for route in routes:
            common_info[f'Valores Limite de Exposicion {route.capitalize()} Valor'] = None
            common_info[f'Valores Limite de Exposicion {route.capitalize()} Unidades'] = None

        valores_limite = hds_dict.get('valoreslimite')
        if not valores_limite:
            return

        for route in routes:
            valor_dict = valores_limite.get(route)
            if valor_dict:
                valor = valor_dict.get('valor')
                unidades = valor_dict.get('unidades')
                try:
                    valor = float(valor)
                except (TypeError, ValueError):
                    valor = None
                common_info[f'Valores Limite de Exposicion {route.capitalize()} Valor'] = valor
                common_info[f'Valores Limite de Exposicion {route.capitalize()} Unidades'] = unidades

    def _merge_cells_by_substance(self, ws, df):
        """Merges cells for columns where the substance is the same."""
        non_merge_columns = set([
            'Nombre del Componente', 'Número CAS del Componente', 'Porcentaje del Componente',
            'Componente GEI', 'Potencial de Calentamiento Global',
            'Componente RETC', 'MPU', 'Emision Transferencia'
        ])


        # print(df.columns)

        columns_to_merge = [col for col in df.columns if col not in non_merge_columns]
        # print(f"Columns to merge: {columns_to_merge}")

        # Initialize tracking for merging
        current_values = {col: None for col in columns_to_merge }
        start_rows = {col: 3 for col in columns_to_merge }  # Data starts from row 3
        current_sustancia = None

        for row in range(3, ws.max_row + 1):
            sustancia_value = ws.cell(row=row, column=2).value  # Column B is 'Nombre de la Sustancia Química'
            if sustancia_value != current_sustancia:
                # New substance encountered
                for col_name in columns_to_merge:
                    # if current_values[col_name] is not None:
                    col_idx = df.columns.get_loc(col_name) + 1
                    # Ensure start_row < end_row before merging
                    if start_rows[col_name] < row - 1:
                        ws.merge_cells(start_row=start_rows[col_name], start_column=col_idx,
                                    end_row=row - 1, end_column=col_idx)
                    current_values[col_name] = ws.cell(row=row, column=df.columns.get_loc(col_name) + 1).value
                    start_rows[col_name] = row
                current_sustancia = sustancia_value
            else:
                # Same substance, clear duplicate values
                for col_name in columns_to_merge :
                    col_idx = df.columns.get_loc(col_name) + 1
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value == current_values[col_name]:
                        ws.cell(row=row, column=col_idx).value = None  # Clear duplicate cell
                    else:
                        if current_values[col_name] is not None:
                            # Ensure start_row < end_row before merging
                            if start_rows[col_name] < row - 1:
                                ws.merge_cells(start_row=start_rows[col_name], start_column=col_idx,
                                            end_row=row - 1, end_column=col_idx)
                        current_values[col_name] = cell_value
                        start_rows[col_name] = row

        # Merge cells for the last group
        for col_name in columns_to_merge:
            if current_values[col_name] is not None:
                col_idx = df.columns.get_loc(col_name) + 1
                # Ensure start_row < end_row before merging
                if start_rows[col_name] < ws.max_row:
                    ws.merge_cells(start_row=start_rows[col_name], start_column=col_idx,
                                end_row=ws.max_row, end_column=col_idx)
                
    def _add_pictogram_images(self, ws, df):
        """
        Replaces pictogram column headers with corresponding images.
        """
    
        pictogram_columns = [
            'Pictogramas Explosivos', 'Pictogramas Inflamables', 'Pictogramas Comburentes', 'Pictogramas Gases Comprimidos',
            'Pictogramas Corrosivos', 'Pictogramas Toxicidad Aguda', 'Pictogramas Irritantes',
            'Pictogramas Peligro Crónico para la Salud', 'Pictogramas Peligro Ambiental'
        ]

        pictogram_col_indices = [
            df.columns.get_loc(col) + 1 for col in pictogram_columns if col in df.columns
        ]

        for col_index in pictogram_col_indices:
            col_letter = get_column_letter(col_index)
            hazard_name = ws.cell(row=2, column=col_index).value

            # Get pictogram image details from the mapping
            pictogram_info = self.peligro_a_pictograma.get(hazard_name)
            if pictogram_info and 'ruta_imagen' in pictogram_info:
                image_path = self.config_path / pictogram_info['ruta_imagen']

                if os.path.exists(image_path):
                    img = Image(image_path)
                    img.height = 45  # Adjust as needed
                    img.width = 45
                    # Center the image in the cell
                    img.anchor = f"{col_letter}2"
                    ws.add_image(img)
                    # ws.column_dimensions[col_letter].width = 15  # Adjust column width as needed
                    # ws.row_dimensions[2].height = 40  # Adjust row height as needed

                    # Clear the text in the header cell
                    ws.cell(row=2, column=col_index).value = None
                else:
                    print(f"No se encontró la imagen para {hazard_name} en {image_path}")
            else:
                print(f"No hay información de pictograma para el peligro: {hazard_name}")
    def _calculate_riesgo_tabla_5(self, common_info: dict) -> None:
        """Calculate 'Grado de Riesgo a la Salud' according to Tabla 5 and add to common_info."""
        grado_oral = None
        grado_cutanea = None
        grado_inhalacion = None

        # Get exposure limit values and units
        oral_val = common_info.get('Valores Limite de Exposicion Oral Valor')
        oral_unit = common_info.get('Valores Limite de Exposicion Oral Unidades')

        cutanea_val = common_info.get('Valores Limite de Exposicion Cutanea Valor')
        cutanea_unit = common_info.get('Valores Limite de Exposicion Cutanea Unidades')

        inhalacion_val = common_info.get('Valores Limite de Exposicion Inhalacion Valor')
        inhalacion_unit = common_info.get('Valores Limite de Exposicion Inhalacion Unidades')

        # Evaluate grado for oral
        if oral_val is not None and oral_unit == 'mg/kg':
            val = oral_val
            if val <= 1:
                grado_oral = 4
            elif val <= 20:
                grado_oral = 3
            elif val <= 50:
                grado_oral = 2
            elif val <= 500:
                grado_oral = 1
            elif val > 5000:
                grado_oral = 0

        # Evaluate grado for cutánea
        if cutanea_val is not None and cutanea_unit == 'mg/kg':
            val = cutanea_val
            if val <= 20:
                grado_cutanea = 4
            elif val <= 200:
                grado_cutanea = 3
            elif val <= 1000:
                grado_cutanea = 2
            elif val <= 5000:
                grado_cutanea = 1
            elif val > 5000:
                grado_cutanea = 0

        # Evaluate grado for inhalación
        if inhalacion_val is not None:
            val = inhalacion_val
            if inhalacion_unit == 'mg/l':
                if val <= 0.2:
                    grado_inhalacion = 4
                elif val <= 2:
                    grado_inhalacion = 3
                elif val <= 20:
                    grado_inhalacion = 2
                elif val <= 200:
                    grado_inhalacion = 1
                elif val > 200:
                    grado_inhalacion = 0
            elif inhalacion_unit == 'ppm':
                if val <= 20:
                    grado_inhalacion = 4
                elif val <= 200:
                    grado_inhalacion = 3
                elif val <= 1000:
                    grado_inhalacion = 2
                elif val <= 10000:
                    grado_inhalacion = 1
                elif val > 10000:
                    grado_inhalacion = 0
            elif inhalacion_unit == 'mg/m³':
                if val <= 200:
                    grado_inhalacion = 4
                elif val <= 2000:
                    grado_inhalacion = 3
                elif val <= 20000:
                    grado_inhalacion = 2
                elif val <= 200000:
                    grado_inhalacion = 1
                elif val > 200000:
                    grado_inhalacion = 0

        # Determine the maximum grado
        grados = [grado for grado in [grado_oral, grado_cutanea, grado_inhalacion] if grado is not None]
        if grados:
            grado_max = max(grados)
        else:
            grado_max = None

        grado_desc = {
            0: 'Grado 0, Mínimamente peligroso',
            1: 'Grado 1, Ligeramente peligroso',
            2: 'Grado 2, Moderadamente peligroso',
            3: 'Grado 3, Seriamente peligroso',
            4: 'Grado 4, Severamente peligroso',
        }

        grado_text = grado_desc.get(grado_max, 'No determinado')

        # Add the result to common_info
        common_info['NOM-010-STPS Riesgo Tabla 5'] = grado_text

    def _calculate_riesgo_tabla_6(self, hds_dict: dict, common_info: dict) -> None:
        """Calculate 'Categoría de Peligro para la Salud' according to Tabla 6 and add to common_info."""
        h_codes = [h['codigo'] for h in hds_dict.get('identificaciones_peligro_h', [])]
        h_codes_set = set(h_codes)
        
        categories = [
            (1, ['H304', 'H330', 'H334', 'H340', 'H350', 'H360', 'H370', 'H372',
                ('H300', 'H330'), ('H310', 'H330'), ('H300', 'H310', 'H330')]),
            (2, ['H305', 'H341', 'H351', 'H361', 'H371', 'H373']),
            (3, ['H331', 'H335', 'H336',
                ('H301', 'H331'), ('H311', 'H331'), ('H301', 'H311', 'H331')]),
            (4, ['H332',
                ('H302', 'H332'), ('H312', 'H332'), ('H302', 'H312', 'H332')]),
            (5, ['H333',
                ('H303', 'H333'), ('H313', 'H333'), ('H303', 'H313', 'H333')]),
        ]
        
        applicable_categories = []

        # Collect all applicable categories and their corresponding H codes
        for category, h_codes_list in categories:
            for h_code_item in h_codes_list:
                if isinstance(h_code_item, str):
                    if h_code_item in h_codes_set:
                        applicable_categories.append((category, [h_code_item]))
                elif isinstance(h_code_item, tuple):
                    if all(h in h_codes_set for h in h_code_item):
                        applicable_categories.append((category, list(h_code_item)))

        if applicable_categories:
            # Select the lowest risk category (highest category number)
            lowest_risk_category = max(applicable_categories, key=lambda x: x[0])
            category_number = lowest_risk_category[0]
            # Collect all H codes corresponding to the lowest risk category
            h_codes_matched = set()
            for cat, codes in applicable_categories:
                if cat == category_number:
                    h_codes_matched.update(codes)
            h_codes_str = ', '.join(sorted(h_codes_matched))
            common_info['NOM-010-STPS Peligro Tabla 6'] = f'Categoría {category_number}, por Peligros H: {h_codes_str}'
        else:
            common_info['NOM-010-STPS Peligro Tabla 6'] = 'No determinado'
    def _determine_volatilidad(self, hds_dict: dict, common_info: dict) -> None:
        """Determine the volatility of the substance and add to common_info."""
        estado_fisico = common_info.get('Estado Físico', '').lower()
        boiling_point = common_info.get('temperatura_ebullicion (Valor)')
        if boiling_point:
            try:
                boiling_point = float(boiling_point)
            except ValueError:
                boiling_point = None

        # Assume operating temperature is 20°C if not provided
        operating_temp = 40  # You may update this if you have actual data

        volatility = 'No determinado'  # Default value

        if estado_fisico == 'gaseoso':
            volatility = 'Alta'
        elif estado_fisico == 'líquido':
            # Apply Tabla 8 logic
            if boiling_point is not None:
                if boiling_point < 50 and 20 <= operating_temp <= 310:
                    volatility = 'Alta'
                elif 50 <= boiling_point <= 150 and 20 <= operating_temp <= 55:
                    volatility = 'Media'
                elif boiling_point > 150 and 20 <= operating_temp <= 55:
                    volatility = 'Baja'
            else:
                volatility = 'Temperatura de ebullicion no disponible'
        elif estado_fisico == 'sólido':
            # Apply Tabla 7 logic
            estado_fisico_data = hds_dict.get('estado_fisico', {})
            if estado_fisico_data.get('solido_baja'):
                volatility = 'Baja'
            elif estado_fisico_data.get('solido_media'):
                volatility = 'Media'
            elif estado_fisico_data.get('solido_alta'):
                volatility = 'Alta'
            else:
                volatility = 'No determinado'

        common_info['NOM-010-STPS Volatilidad Tabla 7 y 8'] = volatility 

    def _customize_worksheet(self, ws, df):
        """Handles custom formatting of the worksheet."""
       
        # Apply wrap text and center alignment to all data rows
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Adjust row heights for better readability
        for row in range(3, ws.max_row + 1):
            ws.row_dimensions[row].height = 25  # Adjust as needed

        # Merge cells based on substance names
        self._merge_cells_by_substance(ws, df)
        
    
    def get_merged_cell_value(self,ws, row, col):
        """
        Gets the value of a merged cell in an Excel worksheet.
        
        Args:
            ws (Worksheet): The Excel worksheet object.
            row (int): Row number of the cell.
            col (int): Column number of the cell.
            
            Returns:
            str: The value of the merged cell.
            
            """
        cell = ws.cell(row=row, column=col)
        if cell.value:
            return cell.value
        else:
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    return ws.cell(row=min_row, column=min_col).value
        return None

    def get_template_headers(self,ws):
        """
        Extracts the headers from the template file.

        Args:
            ws (Worksheet): The Excel worksheet object.

        Returns:
            list: List of header strings.
        """
        headers = []
        max_column = ws.max_column
        for col in range(1, max_column + 1):
            header_parts = []
            # Get header part from row 1
            value_row1 = self.get_merged_cell_value(ws, 1, col)
            if value_row1:
                header_parts.append(value_row1)
            # Get header part from row 2
            value_row2 = self.get_merged_cell_value(ws, 2, col)
            if value_row2 and value_row2 != value_row1:
                header_parts.append(value_row2)
            header = ' '.join(header_parts)
            headers.append(header)
        return headers
    

    def normalize_str(self,s):
        """
        Normalizes a string by removing accents, converting to lowercase, and removing special characters.

        Args:
            s (str): The input string to normalize.

        Returns:
            str: The normalized string.
        """

        if not s:
            return ''
        # Remove accents
        s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
        # Convert to lowercase
        s = s.lower()
        # Remove special characters and spaces
        s = re.sub(r'[\s_()\-\.]', '', s)
        # Remove common Spanish stopwords
        stopwords = ['de', 'del', 'la', 'el', 'los', 'las', 'un', 'una', 'unos', 'unas', 'y', 'o', 'para']
        for word in stopwords:
            s = s.replace(word, '')
        return s

    def export_to_excel_with_template(self, data: list, output_file: str):
        """
        Exports data to an Excel file using a predefined template.

        Args:
            data (list): List of data dictionaries to export.
            template_file (str): Path to the Excel template file.
            output_file (str): Path to save the output Excel file.
        """
        try:
            # Create a DataFrame
            df = pd.DataFrame(data)
            print(f"DataFrame shape: {df.shape}")
            print("DataFrame columns:", df.columns.tolist())


            
             # Define the desired order of columns
            column_order = [
                'Archivo',                         # 1. File name
                'Nombre de la Sustancia Química',   # 2. Substance name
                'Idioma de la HDS',                 # 3. Language
                'Sujeta a GEI',                     # 5. GEI component
                'Sujeta a RETC',
                'Nombre del Componente',            # 4. Component name
                'Número CAS del Componente',        # 7. Component CAS number
                'Porcentaje del Componente',        # 8. Component percentage
                'Componente GEI',                   # 9. GEI component
                'Potencial de Calentamiento Global',# 10. Global Warming Potential
                'Componente RETC',                  # 11. RETC component
                'MPU',                              # 12. MPU
                'Emision Transferencia',            # 13. Emission transfer
                'Palabra de Advertencia',           # 14. Warning word
                'Indicaciones de toxicología',      # 15. Toxicology indications
                'Pictogramas Explosivos',                       # 16. Explosives
                'Pictogramas Inflamables',                      # 17. Flammable
                'Pictogramas Comburentes',                      # 18. Oxidizing
                'Pictogramas Gases Comprimidos',                # 19. Compressed gases
                'Pictogramas Corrosivos',                       # 20. Corrosive
                'Pictogramas Toxicidad Aguda',                  # 21. Acute toxicity
                'Pictogramas Irritantes',                       # 22. Irritants
                'Pictogramas Peligro Crónico para la Salud',    # 23. Chronic health hazard
                'Pictogramas Peligro Ambiental',                # 24. Environmental hazard
                'Estado Físico',                    # 25. Physical state
                'Identificaciones de peligro H',    # 26. Hazard identifications H
                'Consejos de Prudencia P',          # 27. Precautionary statements P
                'Olor',                             # 28. Odor
                'Color',                            # 29. Color
                'pH de la sustancia',               # 30. pH
                'Peso Molecular Valor',             # 31. Molecular weight (Value)
                'Peso Molecular Unidades',          # 32. Molecular weight (Units)
                'temperatura_ebullicion (Valor)',   # 31. Boiling temperature (Value)
                'temperatura_ebullicion (Unidades)',# 32. Boiling temperature (Units)
                'punto_congelacion (Valor)',        # 33. Freezing point (Value)
                'punto_congelacion (Unidades)',     # 34. Freezing point (Units)
                'densidad (Valor)',                 # 35. Density (Value)
                'densidad (Unidades)',              # 36. Density (Units)
                'punto_inflamacion (Valor)',        # 37. Flash point (Value)
                'punto_inflamacion (Unidades)',     # 38. Flash point (Units)
                'Velocidad de Evaporación Valor',    # 39. Evaporation rate (Value)
                'Velocidad de Evaporación Unidades', # 40. Evaporation rate (Units)
                'presion_vapor (Valor)',            # 41. Vapor pressure (Value)
                'presion_vapor (Unidades)',         # 42. Vapor pressure (Units)
                'Solubilidad en Agua Valor',         # 43. Water solubility (Value)
                'Solubilidad en Agua Unidades',      # 44. Water solubility (Units)
                'Propiedades Explosivas',           # 45. Explosive properties
                'Propiedades Comburentes',          # 46. Oxidizing properties
                'Tamaño de Partícula',              # 47. Particle size
                'Límite inferior de inflamabilidad',# 48. Lower flammability limit
                'Límite superior de inflamabilidad', # 49. Upper flammability limit
                'Valores Limite de Exposicion Oral Valor', # 50. Exposure limit values
                'Valores Limite de Exposicion Oral Unidades', # 50. Exposure limit values
                'Valores Limite de Exposicion Inhalacion Valor', # 52. Exposure limit values
                'Valores Limite de Exposicion Inhalacion Unidades', # 52. Exposure limit values
                'Valores Limite de Exposicion Cutanea Valor', # 51. Exposure limit values
                'Valores Limite de Exposicion Cutanea Unidades', # 51. Exposure limit values
                'NOM-010-STPS Riesgo Tabla 5',
                'NOM-010-STPS Peligro Tabla 6',
                'NOM-010-STPS Volatilidad Tabla 7 y 8'  
            ]
            df = df[column_order]  # Ensure columns are in the desired order



            # Load the template
            wb = load_workbook(self.config_path / "template_tabla.xlsx")
            ws = wb.active

            # Get headers from the template
            template_headers = self.get_template_headers(ws)
            # print(f"Template headers: {template_headers}")

            # Create a mapping from DataFrame columns to template columns
            column_mapping = {}
            for idx, header in enumerate(template_headers):
                header_norm = self.normalize_str(header)
                for df_col in df.columns:
                    df_col_norm = self.normalize_str(df_col)
                    if df_col_norm == header_norm:
                        column_mapping[df_col] = idx + 1  # Excel columns are 1-indexed
                        # print(f"Mapping DataFrame column '{df_col}' to template header '{header}'")
                        break
                else:
                    print(f"Header '{header}' in template not found in DataFrame columns.")

            # print(f"Column mapping: {column_mapping}")

            # Start populating data from row 3
            start_row = 3
            for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
                # print(f"Writing data to row {row}")
                for col_name, value in zip(df.columns, row):
                    col_idx = column_mapping.get(col_name)
                    if col_idx:
                        ws.cell(row=row_idx, column=col_idx).value = value
                        # Apply wrap text and alignment
                        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                            wrap_text=True, horizontal='center', vertical='center'
                        )
                    else:
                        print(f"No column mapping found for '{col_name}'")

            # Apply custom formatting and merge logic
            self._customize_worksheet(ws, df)
            self._add_pictogram_images(ws, df)

            # Save the populated file
            wb.save(output_file)
            print(f"Excel file saved to {output_file}")
        except Exception as e:
            print(f"An error occurred: {e}")
            raise e
if __name__ == "__main__":
    # import json 
    # from datetime import datetime
    #Test the class
    generator=  GeneradorTablaSustQ()
    with open("ejemplo/outputs/ejemplo_gral_run_NOM10_raw_data.json", "r", encoding="utf-8") as f:
        hds_data = json.load(f)
    example_flat = generator.flatten_hds_data(hds_data)
    # print(f"Flattened data: {example_flat}")
    generator.export_to_excel_with_template(example_flat,f"test_excel_{str(datetime.timestamp(datetime.now()))}.xlsx")
